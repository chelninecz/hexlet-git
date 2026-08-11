#!/usr/bin/env python3
"""
Генератор Excel файла для игры Блэкджек без VBA.
Все вычисления выполняются формулами Excel.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Color
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils.exceptions import IllegalCharacterError
from openpyxl.workbook.defined_name import DefinedName

def create_blackjack_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "Game"
    
    # Настройка стилей
    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=14, color="000080")
    result_font = Font(bold=True, size=16)
    cell_alignment = Alignment(horizontal='center', vertical='center')
    left_alignment = Alignment(horizontal='left', vertical='center')
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    blue_fill = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
    
    # === ЗАГОЛОВОК ===
    ws.merge_cells('A1:E1')
    ws['A1'] = 'БЛЭКДЖЕК - Игра на формулах Excel'
    ws['A1'].font = title_font
    ws['A1'].alignment = cell_alignment
    
    # === ВХОДНЫЕ ДАННЫЕ ===
    ws['A3'] = 'Seed для новой игры:'
    ws['A3'].font = header_font
    ws['B3'] = 12345  # Начальное значение seed
    ws['B3'].alignment = cell_alignment
    
    ws['A4'] = 'Карт игрока (Hit):'
    ws['A4'].font = header_font
    ws['B4'] = 2  # Начальное количество карт
    ws['B4'].alignment = cell_alignment
    
    ws['A5'] = 'Stand (остановиться):'
    ws['A5'].font = header_font
    ws['B5'] = False
    ws['B5'].alignment = cell_alignment
    
    # Проверка данных для Seed
    dv_seed = DataValidation(type='whole', operator='between', formula1='1', formula2='1000000', allow_blank=False)
    dv_seed.error = 'Введите целое число от 1 до 1000000'
    dv_seed.errorTitle = 'Неверное значение Seed'
    ws.add_data_validation(dv_seed)
    dv_seed.add('B3')
    
    # Проверка данных для количества карт игрока
    dv_player_n = DataValidation(type='whole', operator='between', formula1='2', formula2='20', allow_blank=False)
    dv_player_n.error = 'Введите целое число от 2 до 20'
    dv_player_n.errorTitle = 'Неверное количество карт'
    ws.add_data_validation(dv_player_n)
    dv_player_n.add('B4')
    
    # Проверка данных для Stand (булево)
    dv_stand = DataValidation(type='list', formula1='"FALSE,TRUE"', allow_blank=False)
    dv_stand.error = 'Выберите FALSE или TRUE'
    dv_stand.errorTitle = 'Неверное значение Stand'
    ws.add_data_validation(dv_stand)
    dv_stand.add('B5')
    
    # === РАЗДЕЛИТЕЛЬ ===
    ws['A7'] = '=' * 50
    ws['A7'].font = Font(color="808080")
    
    # === ВЫХОДНЫЕ ДАННЫЕ - ИГРОК ===
    ws['A8'] = '=== РУКА ИГРОКА ==='
    ws['A8'].font = header_font
    
    ws['A9'] = 'Карты игрока:'
    ws['B9'] = ''  # Будет заполнено формулой
    ws['B9'].font = Font(size=14)
    
    ws['A10'] = 'Очки игрока:'
    ws['B10'] = ''  # Будет заполнено формулой
    ws['B10'].font = result_font
    
    # === РАЗДЕЛИТЕЛЬ ===
    ws['A12'] = '=' * 50
    ws['A12'].font = Font(color="808080")
    
    # === ВЫХОДНЫЕ ДАННЫЕ - ДИЛЕР ===
    ws['A13'] = '=== РУКА ДИЛЕРА ==='
    ws['A13'].font = header_font
    
    ws['A14'] = 'Карта дилера (видимая):'
    ws['B14'] = ''  # Будет заполнено формулой
    ws['B14'].font = Font(size=14)
    
    ws['A15'] = 'Карты дилера (после Stand):'
    ws['B15'] = ''  # Будет заполнено формулой
    ws['B15'].font = Font(size=14)
    
    ws['A16'] = 'Очки дилера:'
    ws['B16'] = ''  # Будет заполнено формулой
    ws['B16'].font = result_font
    
    # === РАЗДЕЛИТЕЛЬ ===
    ws['A18'] = '=' * 50
    ws['A18'].font = Font(color="808080")
    
    # === РЕЗУЛЬТАТ ===
    ws['A19'] = '=== РЕЗУЛЬТАТ ==='
    ws['A19'].font = header_font
    
    ws['A20'] = 'Статус игры:'
    ws['B20'] = ''  # Будет заполнено формулой
    ws['B20'].font = result_font
    ws['B20'].alignment = left_alignment
    
    # === ИНСТРУКЦИЯ ===
    ws['A22'] = '=== КАК ИГРАТЬ ==='
    ws['A22'].font = header_font
    
    instructions = [
        '1. Измените Seed (ячейка B3) для начала новой игры',
        '2. Увеличивайте количество карт (B4) чтобы взять карту (Hit)',
        '3. Установите Stand=TRUE (B5) когда решите остановиться',
        '4. Для новой игры: измените Seed и сбросьте карты на 2, Stand на FALSE',
        '',
        'Правила:',
        '- Туз = 11 или 1 (автоматически)',
        '- Карты 2-10 = по номиналу',
        '- J, Q, K = 10',
        '- Дилер берёт до 17 очков',
        '- Победа при > очков чем у дилера (без перебора)'
    ]
    
    for i, instr in enumerate(instructions):
        ws[f'A{23+i}'] = instr
        ws[f'A{23+i}'].font = Font(size=10)
    
    # === СЛУЖЕБНАЯ ОБЛАСТЬ (скрытая логика) ===
    start_col = 'G'
    row_offset = 1
    
    ws[f'{start_col}{row_offset}'] = '=== СЛУЖЕБНАЯ ОБЛАСТЬ ==='
    ws[f'{start_col}{row_offset}'].font = header_font
    row_offset += 2
    
    # Именованные диапазоны будут созданы позже
    # Здесь размещаем вспомогательные вычисления
    
    ws[f'{start_col}{row_offset}'] = 'Deck (колода):'
    ws[f'{get_column_letter(ord(start_col)+1)}{row_offset}'] = ''  # Формула будет добавлена
    row_offset += 1
    
    ws[f'{start_col}{row_offset}'] = 'PlayerCards:'
    ws[f'{get_column_letter(ord(start_col)+1)}{row_offset}'] = ''
    row_offset += 1
    
    ws[f'{start_col}{row_offset}'] = 'DealerCards:'
    ws[f'{get_column_letter(ord(start_col)+1)}{row_offset}'] = ''
    row_offset += 1
    
    ws[f'{start_col}{row_offset}'] = 'PlayerTotal:'
    ws[f'{get_column_letter(ord(start_col)+1)}{row_offset}'] = ''
    row_offset += 1
    
    ws[f'{start_col}{row_offset}'] = 'DealerTotal:'
    ws[f'{get_column_letter(ord(start_col)+1)}{row_offset}'] = ''
    row_offset += 1
    
    ws[f'{start_col}{row_offset}'] = 'Проверка уникальности колоды:'
    ws[f'{get_column_letter(ord(start_col)+1)}{row_offset}'] = ''
    row_offset += 1
    
    ws[f'{start_col}{row_offset}'] = 'Первые 10 карт колоды:'
    row_offset += 1
    for i in range(10):
        ws[f'{get_column_letter(ord(start_col)+i%5)}{row_offset + i//5}'] = f'Карта {i+1}:'
        ws[f'{get_column_letter(ord(start_col)+i%5+1)}{row_offset + i//5}'] = ''
    
    # === СОЗДАНИЕ ИМЕНОВАННЫХ ДИАПАЗОНОВ ===
    # Определяем основные формулы
    
    # Формула для генерации детерминированной колоды
    deck_formula = (
        '=LET('
        'seed, Game!$B$3,'
        'base, SEQUENCE(52),'
        'hash1, MOD((base + seed) * 34999 + 12345, 1048576),'
        'hash2, MOD((base + seed) * 7919 + 104729, 1048576),'
        'SORTBY(base, hash1, 1, hash2, 1)'
        ')'
    )
    
    # Формула PlayerN с защитой от некорректных значений
    player_n_formula = '=MIN(MAX(Game!$B$4, 2), 20)'
    
    # Формула Stand
    stand_formula = '=Game!$B$5'
    
    # Формула карт игрока
    player_cards_formula = (
        '=LET('
        'deck, Deck,'
        'n, PlayerN,'
        'TAKE(deck, n)'
        ')'
    )
    
    # Формула оставшейся колоды после игрока
    remaining_formula = (
        '=LET('
        'deck, Deck,'
        'n, PlayerN,'
        'DROP(deck, n)'
        ')'
    )
    
    # LAMBDA функция для подсчёта очков руки
    hand_total_lambda = (
        '=LAMBDA(cards,'
        'LET('
        'ranks, MOD(cards - 1, 13) + 1,'
        'values, IF(ranks = 1, 11, IF(ranks > 10, 10, ranks)),'
        'total, SUM(values),'
        'aces, SUM(--(ranks = 1)),'
        'reduction, IF(total <= 21, 0, MIN(aces, ROUNDUP((total - 21) / 10, 0))),'
        'total - 10 * reduction'
        ')'
    )
    
    # Формула суммы очков игрока
    player_total_formula = '=HandTotal(PlayerCards)'
    
    # LAMBDA функция для получения названия карты
    card_name_lambda = (
        '=LAMBDA(c,'
        'LET('
        'r, MOD(c - 1, 13) + 1,'
        's, INT((c - 1) / 13) + 1,'
        'rankNames, {"A","2","3","4","5","6","7","8","9","10","J","Q","K"},'
        'suitNames, {"♠","♥","♦","♣"},'
        'INDEX(rankNames, r) & INDEX(suitNames, s)'
        ')'
    )
    
    # LAMBDA функция для отображения руки
    hand_display_lambda = (
        '=LAMBDA(hand,'
        'IF(COUNT(hand) = 0, "", TEXTJOIN("  ", TRUE, MAP(hand, CardName)))'
    )
    
    # Формула отображения руки игрока
    player_display_formula = '=HandDisplay(PlayerCards)'
    
    # Формула для определения количества карт дилера (дилер берёт до 17)
    dealer_n_formula = (
        '=LET('
        'remaining, Remaining,'
        'maxCards, COUNT(remaining),'
        'testCounts, SEQUENCE(maxCards, 1, 1, 1),'
        'totals, MAP(testCounts, LAMBDA(n, HandTotal(TAKE(remaining, n)))),\n'
        'validIdx, FILTER(testCounts, (totals >= 17) + (totals > 21) + (testCounts = maxCards)),\n'
        'IF(COUNT(validIdx) = 0, 2, MIN(validIdx))'
        ')'
    )
    
    # Формула карт дилера
    dealer_cards_formula = (
        '=LET('
        'remaining, Remaining,'
        'n, DealerN,'
        'TAKE(remaining, n)'
        ')'
    )
    
    # Формула суммы очков дилера
    dealer_total_formula = '=HandTotal(DealerCards)'
    
    # Формула проверки натурального блэкджека у игрока
    player_blackjack_formula = (
        '=LET('
        'cards, PlayerCards,'
        'n, COUNT(cards),'
        'IF(n <> 2, FALSE,'
        'LET('
        'r1, MOD(INDEX(cards, 1) - 1, 13) + 1,'
        'r2, MOD(INDEX(cards, 2) - 1, 13) + 1,'
        'v1, IF(r1 = 1, 11, IF(r1 > 10, 10, r1)),'
        'v2, IF(r2 = 1, 11, IF(r2 > 10, 10, r2)),'
        '(v1 + v2 = 21)'
        ')'
        ')'
        ')'
    )
    
    # Формула результата игры
    result_formula = (
        '=LET('
        'pTotal, PlayerTotal,'
        'dTotal, DealerTotal,'
        'stand, Stand,'
        'pBJ, PlayerBlackjack,'
        'IF(pTotal > 21, "Перебор игрока",'
        'IF(stand = FALSE,'
        'IF(pBJ, "Блэкджек! Можно остановиться", "Можно взять ещё карту"),'
        'IFS('
        'pBJ AND (dTotal <> 21 OR COUNT(DealerCards) > 2), "Блэкджек! Вы выиграли",'
        'dTotal > 21, "Дилер перебрал",'
        'pTotal = dTotal, "Ничья",'
        'pTotal > dTotal, "Вы выиграли",'
        'TRUE, "Дилер выиграл"'
        ')'
        ')'
        ')'
    )
    
    # Формула видимой карты дилера (до Stand показываем только первую)
    dealer_visible_formula = (
        '=IF(COUNT(DealerCards) = 0, "", '
        'IF(Stand = FALSE, '
        'CardName(INDEX(DealerCards, 1)) & "  ?", '
        'HandDisplay(DealerCards)'
        ')'
        ')'
    )
    
    # Формула полной руки дилера (для отображения после Stand)
    dealer_full_display_formula = '=IF(Stand, HandDisplay(DealerCards), "")'
    
    # Формула отображения очков дилера (скрыто до Stand)
    dealer_total_display_formula = '=IF(Stand, DealerTotal, "?")'
    
    # Проверка уникальности колоды
    uniqueness_check_formula = '=ROWS(UNIQUE(Deck)) = 52'
    
    # Проверка диапазона карт
    min_card_formula = '=MIN(Deck)'
    max_card_formula = '=MAX(Deck)'
    
    # Проверка отсутствия пересечений
    no_overlap_formula = '=SUM(--(ISNUMBER(MATCH(PlayerCards, DealerCards, 0)))) = 0'
    
    # Добавляем именованные диапазоны в книгу через DefinedName
    row = 3
    col_g = ord(start_col)
    
    def add_named_range(name, cell_ref):
        defined_name = DefinedName(name, attr_text=f"Game!{cell_ref}")
        wb.defined_names[defined_name.name] = defined_name
    
    # Размещаем формулы в служебной области
    ws[f'{get_column_letter(col_g+1)}{row}'] = deck_formula
    add_named_range('Deck', f'${get_column_letter(col_g+1)}${row}')
    row += 1
    
    ws[f'{get_column_letter(col_g+1)}{row}'] = player_n_formula
    add_named_range('PlayerN', f'${get_column_letter(col_g+1)}${row}')
    row += 1
    
    ws[f'{get_column_letter(col_g+1)}{row}'] = stand_formula
    add_named_range('Stand', f'${get_column_letter(col_g+1)}${row}')
    row += 1
    
    ws[f'{get_column_letter(col_g+1)}{row}'] = player_cards_formula
    add_named_range('PlayerCards', f'${get_column_letter(col_g+1)}${row}')
    row += 1
    
    ws[f'{get_column_letter(col_g+1)}{row}'] = remaining_formula
    add_named_range('Remaining', f'${get_column_letter(col_g+1)}${row}')
    row += 1
    
    # LAMBDA функции
    ws[f'{get_column_letter(col_g+1)}{row}'] = hand_total_lambda
    add_named_range('HandTotal', f'${get_column_letter(col_g+1)}${row}')
    row += 1
    
    ws[f'{get_column_letter(col_g+1)}{row}'] = player_total_formula
    add_named_range('PlayerTotal', f'${get_column_letter(col_g+1)}${row}')
    row += 1
    
    ws[f'{get_column_letter(col_g+1)}{row}'] = card_name_lambda
    add_named_range('CardName', f'${get_column_letter(col_g+1)}${row}')
    row += 1
    
    ws[f'{get_column_letter(col_g+1)}{row}'] = hand_display_lambda
    add_named_range('HandDisplay', f'${get_column_letter(col_g+1)}${row}')
    row += 1
    
    ws[f'{get_column_letter(col_g+1)}{row}'] = dealer_n_formula
    add_named_range('DealerN', f'${get_column_letter(col_g+1)}${row}')
    row += 1
    
    ws[f'{get_column_letter(col_g+1)}{row}'] = dealer_cards_formula
    add_named_range('DealerCards', f'${get_column_letter(col_g+1)}${row}')
    row += 1
    
    ws[f'{get_column_letter(col_g+1)}{row}'] = dealer_total_formula
    add_named_range('DealerTotal', f'${get_column_letter(col_g+1)}${row}')
    row += 1
    
    ws[f'{get_column_letter(col_g+1)}{row}'] = player_blackjack_formula
    add_named_range('PlayerBlackjack', f'${get_column_letter(col_g+1)}${row}')
    row += 1
    
    ws[f'{get_column_letter(col_g+1)}{row}'] = result_formula
    add_named_range('Result', f'${get_column_letter(col_g+1)}${row}')
    row += 1
    
    ws[f'{get_column_letter(col_g+1)}{row}'] = dealer_visible_formula
    add_named_range('DealerVisible', f'${get_column_letter(col_g+1)}${row}')
    row += 1
    
    ws[f'{get_column_letter(col_g+1)}{row}'] = dealer_full_display_formula
    add_named_range('DealerFullDisplay', f'${get_column_letter(col_g+1)}${row}')
    row += 1
    
    ws[f'{get_column_letter(col_g+1)}{row}'] = dealer_total_display_formula
    add_named_range('DealerTotalDisplay', f'${get_column_letter(col_g+1)}${row}')
    row += 1
    
    # Проверки
    ws[f'{get_column_letter(col_g+1)}{row}'] = uniqueness_check_formula
    row += 1
    
    ws[f'{get_column_letter(col_g+1)}{row}'] = min_card_formula
    row += 1
    
    ws[f'{get_column_letter(col_g+1)}{row}'] = max_card_formula
    row += 1
    
    ws[f'{get_column_letter(col_g+1)}{row}'] = no_overlap_formula
    row += 1
    
    # Теперь заполняем основные ячейки формулами ссылающимися на именованные диапазоны
    ws['B9'].value = '=PlayerDisplay'
    ws['B10'].value = '=PlayerTotal'
    ws['B14'].value = '=DealerVisible'
    ws['B15'].value = '=DealerFullDisplay'
    ws['B16'].value = '=DealerTotalDisplay'
    ws['B20'].value = '=Result'
    
    # Заполняем служебную область первыми 10 картами
    base_row = 25  # После заголовка "Первые 10 карт колоды:"
    for i in range(10):
        col = chr(ord(start_col) + (i % 5) + 1)
        row_num = base_row + i // 5
        cell = ws[f'{col}{row_num}']
        cell.value = f'=IF(COUNT(Deck)>{i}, CardName(INDEX(Deck, {i+1})), "")'
    
    # Применяем стили
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ord(start_col)+2):
        for cell in row:
            cell.alignment = cell_alignment if cell.column <= 2 else left_alignment
    
    # Условное форматирование для результата
    # Зелёный - победа
    ws.conditional_formatting.add(
        'B20',
        FormulaRule(
            formula=['OR(B20="Вы выиграли", B20="Блэкджек! Вы выиграли", B20="Дилер перебрал")'],
            fill=green_fill
        )
    )
    
    # Красный - поражение/перебор
    ws.conditional_formatting.add(
        'B20',
        FormulaRule(
            formula=['OR(B20="Перебор игрока", B20="Дилер выиграл")'],
            fill=red_fill
        )
    )
    
    # Жёлтый - ничья
    ws.conditional_formatting.add(
        'B20',
        FormulaRule(
            formula=['B20="Ничья"'],
            fill=yellow_fill
        )
    )
    
    # Синий - игра продолжается
    ws.conditional_formatting.add(
        'B20',
        FormulaRule(
            formula=['OR(B20="Можно взять ещё карту", ISNUMBER(SEARCH("Блэкджек", B20)))'],
            fill=blue_fill
        )
    )
    
    # Скрываем служебную колонку
    ws.column_dimensions[start_col].hidden = True
    ws.column_dimensions[chr(ord(start_col)+1)].hidden = True
    ws.column_dimensions[chr(ord(start_col)+2)].hidden = True
    ws.column_dimensions[chr(ord(start_col)+3)].hidden = True
    ws.column_dimensions[chr(ord(start_col)+4)].hidden = True
    ws.column_dimensions[chr(ord(start_col)+5)].hidden = True
    
    # Настраиваем ширину колонок
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    
    # Сохраняем файл
    filename = 'Blackjack_Formulas_Only.xlsx'
    wb.save(filename)
    return filename

if __name__ == '__main__':
    filename = create_blackjack_excel()
    print(f"Файл создан: {filename}")
    print("Откройте файл в Excel 365 или Excel 2021+ для лучшей совместимости.")
