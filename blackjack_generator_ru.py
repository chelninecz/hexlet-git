import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName

# Создаем книгу
wb = Workbook()
ws = wb.active
ws.title = "Game"

# Настройки шрифтов и стилей
header_font = Font(bold=True, size=12)
title_font = Font(bold=True, size=14, color="000000")
result_font = Font(bold=True, size=16)
cell_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
gray_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") # Light Green
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")   # Light Red
yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid") # Light Yellow

# --- 1. Создание интерфейса (Ячейки) ---

# Заголовки
ws['A1'] = "Seed для новой игры"
ws['B1'] = 12345
ws['A2'] = "Карт игрока (Hit)"
ws['B2'] = 2
ws['A3'] = "Stand (Остановиться)"
ws['B3'] = False

# Подписи блоков
ws['A5'] = "--- РЕЗУЛЬТАТЫ ИГРЫ ---"
ws['A6'] = "Рука игрока:"
ws['A7'] = "Очки игрока:"
ws['A8'] = "Рука дилера:"
ws['A9'] = "Очки дилера:"
ws['A10'] = "РЕЗУЛЬТАТ:"

# Ячейки вывода (будут заполнены формулами через именованные диапазоны или прямыми ссылками)
# Для простоты и надежности вставим формулы прямо в ячейки вывода, ссылаясь на именованные формулы
ws['B6'] = "=HandDisplay"       # Отображение карт игрока
ws['B7'] = "=PlayerTotal"       # Сумма игрока
ws['B8'] = "=DealerDisplay"     # Отображение карт дилера (скрыто до Stand)
ws['B9'] = "=DealerTotal"       # Сумма дилера
ws['B10'] = "=GameResult"       # Текст результата

# Отладочная зона (скрыта или ниже)
ws['A12'] = "--- ОТЛАДКА (НЕ МЕНЯТЬ) ---"
ws['A13'] = "Колода (первые 10):"
ws['B13'] = '=TEXTJOIN(", "; TRUE; TAKE(Deck; 10))'
ws['A14'] = "Проверка уникальности:"
ws['B14'] = '=IF(ROWS(UNIQUE(Deck))=52; "OK"; "ERROR")'

# Форматирование
for row in range(1, 15):
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].alignment = Alignment(horizontal='right')
    ws[f'B{row}'].border = cell_border
    ws[f'B{row}'].alignment = Alignment(horizontal='center', vertical='center')

ws['B1'].font = Font(bold=True)
ws['B2'].font = Font(bold=True)
ws['B3'].font = Font(bold=True)
ws['B10'].font = result_font
ws['B10'].alignment = Alignment(horizontal='center')

# Ширина колонок
ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 35

# --- 2. Проверка данных (Data Validation) ---

# Seed: целое число 1 - 1000000
dv_seed = DataValidation(type="whole", operator="between", formula1="1", formula2="1000000", allow_blank=False)
dv_seed.error = "Введите число от 1 до 1000000"
ws.add_data_validation(dv_seed)
dv_seed.add('B1')

# Карты игрока: целое число 2 - 20
dv_cards = DataValidation(type="whole", operator="between", formula1="2", formula2="20", allow_blank=False)
dv_cards.error = "Введите число от 2 до 20"
ws.add_data_validation(dv_cards)
dv_cards.add('B2')

# Stand: Список FALSE, TRUE (или логическое)
# В русском Excel список работает надежнее для эмуляции переключателя
dv_stand = DataValidation(type="list", formula1='"FALSE,TRUE"', allow_blank=False)
dv_stand.error = "Выберите FALSE или TRUE"
ws.add_data_validation(dv_stand)
dv_stand.add('B3')

# --- 3. Именованные диапазоны и Формулы (Global Names) ---
# Важно: Формулы пишутся на английском синтаксисе (OOXML стандарт)
# Excel при открытии в русской локали сам переведет функции (IF -> ЕСЛИ, LAMBDA -> ЛЯМБДА)

def add_defined_name(name, formula, local_sheet=None):
    """Создает именованный диапазон"""
    defn = DefinedName(name, attr_text=formula)
    if local_sheet:
        defn.localSheetId = local_sheet
    wb.defined_names.add(defn)

# Вспомогательные константы
add_defined_name("SeedVal", "=Game!$B$1")
add_defined_name("PlayerNInput", "=Game!$B$2")
add_defined_name("StandInput", "=Game!$B$3")

# Ограниченное количество карт игрока (мин 2, макс 20)
add_defined_name("PlayerN", "=MIN(MAX(Game!$B$2; 2); 20)")

# ГЕНЕРАЦИЯ КОЛОДЫ (ДЕТЕРМИНИРОВАННАЯ)
# Используем SORTBY с математической перестановкой на основе Seed
# Формула создает массив 1..52 и сортирует его по псевдо-случайным ключам
deck_formula = (
    "=LET("
    "s; SeedVal;"
    "raw; SEQUENCE(52);"
    "key1; MOD((raw + s) * 34999 + 12345; 1048576);"
    "key2; MOD((raw + s) * 7919 + 104729; 1048576);"
    "SORTBY(raw; key1; 1; key2; 1)"
    ")"
)
add_defined_name("Deck", deck_formula)

# КАРТЫ ИГРОКА
add_defined_name("PlayerCards", "=TAKE(Deck; PlayerN)")

# ОСТАТОК КОЛОДЫ
add_defined_name("RemainingDeck", "=DROP(Deck; PlayerN)")

# ФУНКЦИЯ ПОДСЧЕТА ОЧКОВ (LAMBDA)
# Принимает массив карт, возвращает сумму
# Логика: Туз=11, если >21, уменьшаем на 10 за каждый туз пока возможно
hand_total_formula = (
    "=LAMBDA(cards;"
    "LET("
    "ranks; MOD(cards - 1; 13) + 1;"
    "vals; IF(ranks = 1; 11; IF(ranks > 10; 10; ranks));"
    "sum_raw; SUM(vals);"
    "aces; COUNTIF(ranks; 1);"
    "reduce_by; MIN(aces; MAX(0; ROUNDUP((sum_raw - 21) / 10; 0)));"
    "sum_raw - reduce_by * 10"
    ")"
    ")"
)
add_defined_name("HandTotalFunc", hand_total_formula)

# СЧЕТ ИГРОКА
add_defined_name("PlayerTotal", "=HandTotalFunc(PlayerCards)")

# ЛОГИКА ДИЛЕРА (Сколько карт брать?)
# Дилер берет, пока сумма < 17.
# Реализуем через явный перебор вариантов (стабильнее чем MAP в старых версиях)
dealer_n_formula = (
    "=LET("
    "rem; RemainingDeck;"
    "c2; TAKE(rem; 2); t2; HandTotalFunc(c2);"
    "c3; TAKE(rem; 3); t3; HandTotalFunc(c3);"
    "c4; TAKE(rem; 4); t4; HandTotalFunc(c4);"
    "c5; TAKE(rem; 5); t5; HandTotalFunc(c5);"
    "c6; TAKE(rem; 6); t6; HandTotalFunc(c6);"
    "c7; TAKE(rem; 7); t7; HandTotalFunc(c7);"
    "c8; TAKE(rem; 8); t8; HandTotalFunc(c8);"
    "c9; TAKE(rem; 9); t9; HandTotalFunc(c9);"
    "c10; TAKE(rem; 10); t10; HandTotalFunc(c10);"
    "IF(t2>=17; 2;"
    "IF(t3>=17; 3;"
    "IF(t4>=17; 4;"
    "IF(t5>=17; 5;"
    "IF(t6>=17; 6;"
    "IF(t7>=17; 7;"
    "IF(t8>=17; 8;"
    "IF(t9>=17; 9;"
    "10)))))))))"
    ")"
)
add_defined_name("DealerN", dealer_n_formula)

# КАРТЫ ДИЛЕРА
add_defined_name("DealerCards", "=TAKE(RemainingDeck; DealerN)")

# СЧЕТ ДИЛЕРА
add_defined_name("DealerTotal", "=HandTotalFunc(DealerCards)")

# ОТОБРАЖЕНИЕ ОДНОЙ КАРТЫ (LAMBDA)
# Возвращает строку вида "A♠", "10♥"
card_name_formula = (
    "=LAMBDA(c;"
    "LET("
    "r; MOD(c - 1; 13) + 1;"
    "s; INT((c - 1) / 13) + 1;"
    'rankTxt; CHOOSE(r; "A"; "2"; "3"; "4"; "5"; "6"; "7"; "8"; "9"; "10"; "J"; "Q"; "K");'
    'suitTxt; CHOOSE(s; "♠"; "♥"; "♦"; "♣");'
    "rankTxt & suitTxt"
    ")"
    ")"
)
add_defined_name("CardNameFunc", card_name_formula)

# ОТОБРАЖЕНИЕ РУКИ ИГРОКА
add_defined_name("HandDisplay", '=TEXTJOIN("  "; TRUE; MAP(PlayerCards; CardNameFunc))')

# ОТОБРАЖЕНИЕ РУКИ ДИЛЕРА (С учетом скрытой карты)
# Если Stand=FALSE, показываем первую карту и "?"
# Если Stand=TRUE, показываем все
dealer_display_formula = (
    "=IF(StandInput;"
    'TEXTJOIN("  "; TRUE; MAP(DealerCards; CardNameFunc));'
    'CardNameFunc(INDEX(DealerCards; 1)) & "  ?"'
    ")"
)
add_defined_name("DealerDisplay", dealer_display_formula)

# РЕЗУЛЬТАТ ИГРЫ
# Логика:
# 1. Если PlayerTotal > 21 -> Перебор
# 2. Если Stand=FALSE -> Игра идет
# 3. Если DealerTotal > 21 -> Дилер перебрал
# 4. Если Player > Dealer -> Победа
# 5. Если Player = Dealer -> Ничья
# 6. Иначе -> Дилер выиграл
result_formula = (
    "=LET("
    "p; PlayerTotal;"
    "d; DealerTotal;"
    "stand; StandInput;"
    'IF(p > 21; "Перебор игрока";'
    'IF(stand = FALSE; "Можно взять ещё карту";'
    "IFS("
    'd > 21; "Дилер перебрал";'
    'p > d; "Вы выиграли";'
    'p = d; "Ничья";'
    'TRUE; "Дилер выиграл"'
    ")"
    "))"
    ")"
)
add_defined_name("GameResult", result_formula)

# --- 4. Условное форматирование ---

# Правило: Перебор игрока (Красный)
# Ссылка на ячейку B10 (Результат)
cf_bust = FormulaRule(formula=['=$B$10="Перебор игрока"'], fill=red_fill)
ws.conditional_formatting.add('B10', cf_bust)

# Правило: Победа (Зеленый)
cf_win = FormulaRule(formula=['OR($B$10="Вы выиграли"; $B$10="Дилер перебрал")'], fill=green_fill)
ws.conditional_formatting.add('B10', cf_win)

# Правило: Ничья (Желтый)
cf_push = FormulaRule(formula=['=$B$10="Ничья"'], fill=yellow_fill)
ws.conditional_formatting.add('B10', cf_push)

# Правило: Игра идет (Серый)
cf_play = FormulaRule(formula=['=$B$10="Можно взять ещё карту"'], fill=gray_fill)
ws.conditional_formatting.add('B10', cf_play)

# Сохранение
filename = "Blackjack_RU_Locales.xlsx"
wb.save(filename)
print(f"Файл '{filename}' успешно создан.")
print("Откройте его в Excel. Формулы автоматически отобразятся на русском языке.")
